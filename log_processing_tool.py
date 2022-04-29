# coding=utf-8
import os
import json
import shutil
from datetime import date, timedelta
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import colors

# 配置数据
cfg = {'name': '', 'log_url': '', 'log_file_dir': '', 'start_day': '', 'end_day': ''}
# 存储日志的excel名称
log_excel_name = 'error_log.xlsx'


class LogTool:
    # 日志列表
    log_list = []
    # 日志字典
    log_dic = {}
    # 日志总条数
    log_all_num = 0

    # 解析某天日志文件为字典
    def parse_log(self, log_file_name):
        global cfg
        day_log_dic = {}
        with open(os.path.join(cfg['log_file_dir'], log_file_name), 'r', encoding="utf-8") as file:
            lines = file.readlines()
            self.log_all_num = len(lines)
            for line in lines:
                if not (' - ' in line):
                    continue
                json_str = line.split(' - ')[1]
                json_dic = json.loads(json_str)
                have_error_stack = ("stack" in json_dic) and (json_dic["stack"] != "")
                if not have_error_stack:
                    continue
                error_key = str(json_dic['lineno']) + '_' + str(json_dic['colno'])
                if not (error_key in day_log_dic):
                    day_log_dic[error_key] = {'count': 0, 'rate': 0, 'stack': json_dic['stack']}
                day_log_dic[error_key]['count'] += 1
        for log in day_log_dic.values():
            log['rate'] = str(round(100 * log['count'] / self.log_all_num, 2)) + '%'
        return day_log_dic

    # 将日志字典转为列表并排序
    @staticmethod
    def sort_log(log_map):
        logs = []
        for log in log_map.values():
            logs.append(log)
        logs = sorted(logs, key=lambda error_log: error_log['count'], reverse=True)
        return logs

    # 将日志写入excel
    @staticmethod
    def write_to_excel(work_book, sheet_name, logs):
        # 清空分页并设置表头
        if sheet_name in work_book.sheetnames:
            work_book.remove(work_book[sheet_name])
        sheet = work_book.create_sheet(sheet_name)
        sheet.cell(1, 1).value = '索引'
        sheet.cell(1, 2).value = '计数'
        sheet.cell(1, 3).value = '占比'
        sheet.cell(1, 4).value = '堆栈'

        # 写入数据
        for i, log in enumerate(logs):
            row = i + 2
            sheet.cell(row, 1).value = i + 1
            sheet.cell(row, 2).value = log['count']
            sheet.cell(row, 3).value = log['rate']
            sheet.cell(row, 4).value = log['stack']

        #  美化, 给计数加上蓝色数据条
        sheet.column_dimensions['B'].width = 20
        start_cell = 'B2'
        end_cell = 'B' + str(max(2, sheet.max_row))
        rule = DataBarRule(start_type='percentile',
                           start_value=0,
                           end_type='percentile',
                           end_value=100,
                           color=colors.BLUE)
        sheet.conditional_formatting.add("%s:%s" % (start_cell, end_cell), rule)

    # 获取空工作表
    @staticmethod
    def clear_excel():
        if os.path.isfile(log_excel_name):
            print('===> 如若处理失败, 请检查%s表是否被占用' % log_excel_name)
            os.remove(log_excel_name)
        wb = Workbook(log_excel_name)
        wb.save(log_excel_name)

    # 按天写入数据
    def write_day_log(self):
        wb = load_workbook(log_excel_name)
        for log_file_name in os.listdir(cfg['log_file_dir']):
            day_str = log_file_name.split('.')[0]
            # 解析日志
            day_log_dic = self.parse_log(log_file_name)
            # 日志排序
            day_log_list = self.sort_log(day_log_dic)
            # 写入excel
            self.write_to_excel(wb, day_str, day_log_list)
            print('===> 写入excel页签%s完成' % day_str)
            # 汇总
            for error_key, log in day_log_dic.items():
                if not (error_key in self.log_dic):
                    self.log_dic[error_key] = {'count': 0, 'rate': 0, 'stack': log['stack']}
                self.log_dic[error_key]['count'] += 1
        wb.save(log_excel_name)

    # 写入汇总数据
    def write_all_log(self):
        wb = load_workbook(log_excel_name)
        self.log_all_num = len(self.log_dic.keys())
        for log in self.log_dic.values():
            log['rate'] = str(round(100 * log['count'] / self.log_all_num, 2)) + '%'
        self.log_list = self.sort_log(self.log_dic)
        self.write_to_excel(wb, "全部", self.log_list)
        wb.active = wb['全部']
        print('===> 写入excel页签全部完成')
        print('===> 如若处理失败, 请检查%s表是否被占用' % log_excel_name)
        wb.save(log_excel_name)

    # 处理日志
    def __init__(self):
        # 获取空工作表
        self.clear_excel()
        # 按天写入数据
        self.write_day_log()
        # 写入汇总数据
        self.write_all_log()


class DownloadTool:
    # 获取事件区间列表
    @staticmethod
    def get_day_list():
        day_list: list[date] = []
        start_arr = cfg['start_day'].split('-')
        start_day = date(int(start_arr[0]), int(start_arr[1]), int(start_arr[2]))
        end_arr = cfg['end_day'].split('-')
        end_day = date(int(end_arr[0]), int(end_arr[1]), int(end_arr[2]))
        time_delta = timedelta(days=1)
        cur_day = start_day
        while cur_day <= end_day:
            day_list.append(cur_day)
            cur_day += time_delta
        return day_list

    # 下载某天日志
    @staticmethod
    def download_log(day_str):
        resp = requests.get(cfg['log_url'] % day_str)
        if not os.path.isdir(cfg['log_file_dir']):
            os.mkdir(cfg['log_file_dir'])
        with open(os.path.join(cfg['log_file_dir'], str(day_str) + '.log'), 'wb') as file:
            file.write(resp.content)
            print("===> %s日志下载保存成功" % day_str)

    def __init__(self):
        shutil.rmtree(cfg['log_file_dir'])
        print('===> 日志目录已清除')
        day_list: list[date] = self.get_day_list()
        for day in day_list:
            self.download_log(day)
        print("===> %s--%s全部日志下载保存成功" % (cfg['start_day'], cfg['end_day']))


if __name__ == "__main__":
    with open('./config.json', 'r', encoding='utf-8') as f:
        json_data = json.load(f)
        channel_cfg = json_data['channel_map'][str(json_data['select_channel'])]
        cfg['name'] = channel_cfg['name']
        cfg['log_url'] = channel_cfg['log_url']
        cfg['log_file_dir'] = './log/' + channel_cfg['name']
        cfg['start_day'] = json_data['start_day']
        cfg['end_day'] = json_data['end_day']

    # 执行下载
    DownloadTool()
    # 执行写入
    LogTool()
    print('===> SUCCESS!!! 执行成功!!!')
    pass
